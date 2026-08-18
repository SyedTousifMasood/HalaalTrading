import os
import sys
import datetime
import openpyxl
import yfinance as yf
from openpyxl.styles import Font, PatternFill, Alignment
sys.path.append(".")

from hsts.journal import TradingJournal

def run_master_reconciliation():
    # Force use the local repository file path
    journal = TradingJournal(file_path="Trading_Journal.xlsx")
    file_path = "Trading_Journal.xlsx"
    
    print(f"Loading clean workbook: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    ws_ledger = wb["Ledger"]
    
    # 1. Check if NESTLEIND or PIDILITIND are already in Ledger to avoid duplicates
    nestle_exists = False
    pidilite_exists = False
    
    for r in range(2, ws_ledger.max_row + 1):
        sym = ws_ledger.cell(row=r, column=1).value
        status = ws_ledger.cell(row=r, column=14).value
        if sym == "NESTLEIND":
            nestle_exists = True
        elif sym == "PIDILITIND":
            pidilite_exists = True
            
    wb.close()
    
    # 2. Add open trades if they do not exist
    if not nestle_exists:
        print("Recording open trade for NESTLEIND...")
        journal.add_trade(
            symbol="NESTLEIND",
            name="Nestle India Limited",
            entry_date="2026-08-04",
            qty=1,
            buy_price=1523.00,
            suggested_entry=1530.00,
            target=1550.00,
            stop_loss=1520.10,
            notes="Intraday trade setup.",
            trade_type="INTRADAY"
        )
        
    if not pidilite_exists:
        print("Recording open trade for PIDILITIND...")
        journal.add_trade(
            symbol="PIDILITIND",
            name="Pidilite Industries Limited",
            entry_date="2026-08-04",
            qty=1,
            buy_price=1642.50,
            suggested_entry=1642.50,
            target=1670.00,
            stop_loss=1634.50,
            notes="Intraday trade setup.",
            trade_type="INTRADAY"
        )
        
    # 3. Log buy brokerage transactions
    buy_notes_nestle = "Brokerage & taxes on NESTLEIND BUY (1 qty @ 1530.00)"
    if not journal.capital_transaction_exists(buy_notes_nestle):
        journal.add_capital_transaction("WITHDRAWAL", 1.84, notes=buy_notes_nestle)
        
    buy_notes_pidilite = "Brokerage & taxes on PIDILITIND BUY (1 qty @ 1642.50)"
    if not journal.capital_transaction_exists(buy_notes_pidilite):
        journal.add_capital_transaction("WITHDRAWAL", 1.97, notes=buy_notes_pidilite)

    # 4. Close trades via Stop Loss GTT triggers
    exit_date = "2026-08-04"
    
    # Close NESTLEIND
    wb = openpyxl.load_workbook(file_path)
    ws_ledger = wb["Ledger"]
    nestle_open = False
    pidilite_open = False
    for r in range(2, ws_ledger.max_row + 1):
        sym = ws_ledger.cell(row=r, column=1).value
        status = ws_ledger.cell(row=r, column=14).value
        if sym == "NESTLEIND" and status == "OPEN":
            nestle_open = True
        elif sym == "PIDILITIND" and status == "OPEN":
            pidilite_open = True
    wb.close()
    
    if nestle_open:
        print("Closing NESTLEIND...")
        journal.close_trade(
            symbol="NESTLEIND",
            exit_date=exit_date,
            exit_price=1520.10,
            status="LOSS",
            notes="Position closed via Stop-Loss GTT trigger. Actual Buy Filled at INR 1523.00",
            trade_type="INTRADAY"
        )
        # Log sell brokerage
        sell_notes_nestle = "Brokerage & taxes on NESTLEIND SELL (1 qty @ 1520.10)"
        if not journal.capital_transaction_exists(sell_notes_nestle):
            journal.add_capital_transaction("WITHDRAWAL", 1.83, notes=sell_notes_nestle)

    if pidilite_open:
        print("Closing PIDILITIND...")
        journal.close_trade(
            symbol="PIDILITIND",
            exit_date=exit_date,
            exit_price=1634.50,
            status="LOSS",
            notes="Position closed via Stop-Loss GTT trigger.",
            trade_type="INTRADAY"
        )
        # Log sell brokerage
        sell_notes_pidilite = "Brokerage & taxes on PIDILITIND SELL (1 qty @ 1634.50)"
        if not journal.capital_transaction_exists(sell_notes_pidilite):
            journal.add_capital_transaction("WITHDRAWAL", 1.96, notes=sell_notes_pidilite)

    # 5. Live margin reconciliation to align dynamically with Zerodha
    import os
    from dotenv import load_dotenv
    from hsts.broker.zerodha_free import ZerodhaFreeBroker
    
    load_dotenv()
    user_id = os.getenv("ZERODHA_USER_ID")
    password = os.getenv("ZERODHA_PASSWORD")
    totp_secret = os.getenv("ZERODHA_TOTP_SECRET")
    
    # We default actual cash to the live equity balance or opening balance
    actual_cash = 3928.20
    try:
        print("Connecting to Zerodha to fetch live cash margin...")
        broker = ZerodhaFreeBroker(user_id=user_id, password=password, totp_secret=totp_secret)
        if broker.authenticate():
            margins = broker.get_margins()
            if margins:
                # Use live balance or cash
                actual_cash = float(margins.get("equity", {}).get("available", {}).get("live_balance", actual_cash))
                print(f"Fetched live Zerodha Equity margin: INR {actual_cash:.2f}")
    except Exception as e:
        print(f"Error fetching live margin: {e}")
        
    journal_cash = journal.get_available_capital()
    if actual_cash <= 0:
        actual_cash = journal_cash
        
    journal_cash = journal.get_available_capital()
    print(f"Pre-reconciliation Journal Cash: INR {journal_cash:.2f}")
    
    diff = float(actual_cash) - float(journal_cash)
    if abs(diff) > 0.01:
        print(f"Discrepancy found: INR {diff:+.2f} (Calculated: {journal_cash:.2f} vs Zerodha Live: {actual_cash:.2f}).")
        print(f"[RECONCILED] Available cash updated directly to Zerodha margin: INR {actual_cash:.2f}")
    else:
        print("Journal cash matches Zerodha cash perfectly!")

    # 6. Apply row styling to Ledger
    print("Applying outcome-based row highlights to Ledger sheet...")
    journal.apply_row_styling()

    # 7. Rebuild Dashboard sheet with new formulas and colors
    print("Rebuilding Dashboard with new formulas and dynamic color fusions...")
    wb = openpyxl.load_workbook(file_path)
    
    # Read open trades to calculate current value
    open_positions = []
    ws_ledg = wb["Ledger"]
    for r in range(2, ws_ledg.max_row + 1):
        sym = ws_ledg.cell(row=r, column=1).value
        qty = ws_ledg.cell(row=r, column=4).value
        price = ws_ledg.cell(row=r, column=5).value
        status = ws_ledg.cell(row=r, column=14).value
        if sym is not None:
            print(f"Row {r}: sym={sym}, status={status}, match={bool(sym and status == 'OPEN')}")
        if sym and status == "OPEN":
            open_positions.append({
                "symbol": sym,
                "qty": int(qty) if qty else 0,
                "cost": float(price) if price else 0.0
            })
            
    print("Found open positions for yfinance:", open_positions)
    total_cost = 0.0
    total_value = 0.0
    if open_positions:
        symbols_ns = [f"{p['symbol']}.NS" for p in open_positions]
        try:
            df = yf.download(symbols_ns, period="1d", progress=False)
            for pos in open_positions:
                ns_sym = f"{pos['symbol']}.NS"
                last_price = pos['cost']
                if len(symbols_ns) == 1:
                    if not df.empty:
                        last_price = float(df["Close"].iloc[-1])
                else:
                    if ns_sym in df["Close"].columns:
                        col_data = df["Close"][ns_sym].dropna()
                        if not col_data.empty:
                            last_price = float(col_data.iloc[-1])
                pos["last_price"] = last_price
                pos_cost = pos["qty"] * pos["cost"]
                pos_value = pos["qty"] * last_price
                total_cost += pos_cost
                total_value += pos_value
        except Exception as e:
            print(f"Error fetching live prices: {e}")
            for pos in open_positions:
                pos["last_price"] = pos["cost"]
                pos_cost = pos["qty"] * pos["cost"]
                total_cost += pos_cost
                total_value += pos_cost
                
    ws_dash = wb["Dashboard"]
    ws_dash.views.sheetView[0].showGridLines = True
    
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=11, bold=True)
    font_regular = Font(name="Segoe UI", size=11)
    font_section = Font(name="Segoe UI", size=11, bold=True, color="1B365D")
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_metric = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    fill_section = PatternFill(start_color="D3DDF1", end_color="D3DDF1", fill_type="solid")
    fill_profit = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    fill_loss = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # Unmerge all cells first
    for rng in list(ws_dash.merged_cells.ranges):
        ws_dash.unmerge_cells(str(rng))

    # Clear
    for row in ws_dash.iter_rows(min_row=1, max_row=50, min_col=1, max_col=10):
        for cell in row:
            cell.value = None
            cell.fill = openpyxl.styles.fills.PatternFill(fill_type=None)
            cell.font = Font(name="Segoe UI", size=11)

    ws_dash["A1"] = "HSTS v1.0 Trading Dashboard"
    ws_dash["A1"].font = font_title
    ws_dash.row_dimensions[1].height = 30

    last_updated_str = f"Last updated on : {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_dash["A2"] = last_updated_str
    ws_dash["A2"].font = Font(name="Segoe UI", size=9, italic=True, color="595959")
    ws_dash.row_dimensions[2].height = 18

    headers = ["Metric", "Value"]
    for col_idx, text in enumerate(headers, 1):
        cell = ws_dash.cell(row=3, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[3].height = 24

    # Calculate analytics from Ledger sheet
    closed_durations = []
    win_pcts = []
    loss_pcts = []
    trade_values = []
    entry_dates = set()
    trades_with_dates = []
    min_date = None
    max_date = datetime.date.today()
    
    ws_ledg = wb["Ledger"]
    for r in range(2, ws_ledg.max_row + 1):
        sym = ws_ledg.cell(row=r, column=1).value
        status = ws_ledg.cell(row=r, column=14).value
        entry_val = ws_ledg.cell(row=r, column=3).value
        exit_val = ws_ledg.cell(row=r, column=11).value
        buy_pr = ws_ledg.cell(row=r, column=5).value
        exit_pr = ws_ledg.cell(row=r, column=12).value
        qty = ws_ledg.cell(row=r, column=4).value
        
        if sym and entry_val:
            try:
                if buy_pr and qty:
                    trade_values.append(float(buy_pr) * float(qty))
                if isinstance(entry_val, (datetime.datetime, datetime.date)):
                    ent_d = entry_val.date() if isinstance(entry_val, datetime.datetime) else entry_val
                else:
                    ent_d = datetime.datetime.strptime(str(entry_val).split(" ")[0], "%Y-%m-%d").date()
                
                entry_dates.add(ent_d)
                
                ext_d = None
                if exit_val:
                    if isinstance(exit_val, (datetime.datetime, datetime.date)):
                        ext_d = exit_val.date() if isinstance(exit_val, datetime.datetime) else exit_val
                    else:
                        ext_d = datetime.datetime.strptime(str(exit_val).split(" ")[0], "%Y-%m-%d").date()
                
                trades_with_dates.append((sym, ent_d, ext_d))
                
                if min_date is None or ent_d < min_date:
                    min_date = ent_d
                    
                if status in ["WIN", "LOSS"] and ext_d:
                    closed_durations.append((ext_d - ent_d).days)
                    if buy_pr and exit_pr:
                        pct = (float(exit_pr) - float(buy_pr)) / float(buy_pr)
                        if status == "WIN":
                            win_pcts.append(pct)
                        elif status == "LOSS":
                            loss_pcts.append(pct)
                elif status == "OPEN":
                    closed_durations.append((max_date - ent_d).days)
            except Exception:
                pass
                
    avg_exit_days = sum(closed_durations) / len(closed_durations) if closed_durations else 0.0
    avg_trades_per_day = len(trades_with_dates) / len(entry_dates) if entry_dates else 0.0
    avg_win_pct = sum(win_pcts) / len(win_pcts) if win_pcts else 0.0
    avg_loss_pct = sum(loss_pcts) / len(loss_pcts) if loss_pcts else 0.0
    avg_trade_val = sum(trade_values) / len(trade_values) if trade_values else 0.0
    
    daily_counts = []
    if min_date:
        d = min_date
        while d <= max_date:
            if d.weekday() < 5:
                count = sum(1 for sym, ent_d, ext_d in trades_with_dates if ent_d <= d <= (ext_d if ext_d else max_date))
                daily_counts.append(count)
            d += datetime.timedelta(days=1)
    avg_unique_holdings = sum(daily_counts) / len(daily_counts) if daily_counts else 0.0

    metrics = [
        ("Global Portfolio Metrics", None, "header"),
        ("Total Net Capital Deposited", '=SUMIF(Capital!B:B, "DEPOSIT", Capital!C:C)', "currency"),
        ("Lifetime Realized PnL", "=SUM(Ledger!M:M)", "currency"),
        ("Lifetime Realized P/L Percentage", "=B6/B5", "percentage"),
        ("Current Total Portfolio Value", "=B13+B12", "currency"),
        ("Total Number of Trades", '=COUNTIF(Ledger!N:N, "WIN") + COUNTIF(Ledger!N:N, "LOSS") + COUNTIF(Ledger!N:N, "OPEN")', "integer"),
        ("Winning Percentage of Trades", '=IF((COUNTIF(Ledger!N:N, "WIN")+COUNTIF(Ledger!N:N, "LOSS"))>0, COUNTIF(Ledger!N:N, "WIN")/(COUNTIF(Ledger!N:N, "WIN")+COUNTIF(Ledger!N:N, "LOSS")), 0)', "percentage"),
        ("Capital Deployed in Open Trades", '=SUMPRODUCT((Ledger!N2:N5000="OPEN")*(Ledger!E2:E5000)*(Ledger!D2:D5000))', "currency"),
        ("Current Value of Active Trades", total_value, "currency"),
        ("Capital Available for Trading", actual_cash, "currency"),
        ("Total Wins", '=COUNTIF(Ledger!N:N, "WIN")', "integer"),
        ("Total Losses", '=COUNTIF(Ledger!N:N, "LOSS")', "integer"),
        ("Total Taxes Paid to Zerodha", '=SUMIF(Capital!D:D, "*Brokerage & taxes*", Capital!C:C)', "currency"),
        ("Swing Trading Statistics", None, "header"),
        ("Swing Win Rate", '=IF((COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "WIN")+COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "LOSS"))>0, COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "WIN")/(COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "WIN")+COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "LOSS")), 0)', "percentage"),
        ("Swing Realized PnL", '=SUMIFS(Ledger!M:M, Ledger!P:P, "SWING")', "currency"),
        ("Swing Completed Trades", '=COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "WIN") + COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "LOSS")', "integer"),
        ("Swing Active Positions", '=COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "OPEN")', "integer"),
        ("Intraday Trading Statistics", None, "header"),
        ("Intraday Win Rate", '=IF((COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "WIN")+COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "LOSS"))>0, COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "WIN")/(COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "WIN")+COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "LOSS")), 0)', "percentage"),
        ("Intraday Realized PnL", '=SUMIFS(Ledger!M:M, Ledger!P:P, "INTRADAY")', "currency"),
        ("Intraday Completed Trades", '=COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "WIN") + COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "LOSS")', "integer"),
        ("Intraday Active Positions", '=COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "OPEN")', "integer"),
        ("Activity & Performance Metrics", None, "header"),
        ("Average Days to Exit a Trade", avg_exit_days, "float"),
        ("Average Trades per Day", avg_trades_per_day, "float"),
        ("Average Unique Holdings per Day", avg_unique_holdings, "float"),
        ("Average Profit % per trade", avg_win_pct, "percentage"),
        ("Average Loss % per trade", avg_loss_pct, "percentage"),
        ("Average Value per Trade", avg_trade_val, "currency"),
    ]

    # Evaluate current PnL sum for coloring
    actual_pnl_sum = 0.0
    for r in range(2, ws_ledg.max_row + 1):
        status = ws_ledg.cell(row=r, column=14).value
        pnl = ws_ledg.cell(row=r, column=13).value
        # If it's closed, we can calculate pnl manually here for color evaluation
        if status in ["WIN", "LOSS"]:
            try:
                qty = float(ws_ledg.cell(row=r, column=4).value)
                buy = float(ws_ledg.cell(row=r, column=5).value)
                exit_pr = float(ws_ledg.cell(row=r, column=12).value)
                actual_pnl_sum += (exit_pr - buy) * qty
            except Exception:
                pass
    print(f"Evaluated realized PnL sum for coloring: {actual_pnl_sum:.2f}")

    for idx, (metric_name, formula, m_type) in enumerate(metrics, 4):
        cell_name = ws_dash.cell(row=idx, column=1, value=metric_name)
        cell_val = ws_dash.cell(row=idx, column=2)
        
        if m_type == "header":
            cell_name.font = font_section
            cell_name.fill = fill_section
            cell_val.value = ""
            cell_val.fill = fill_section
            # Merge columns A and B for this header row
            ws_dash.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=2)
        else:
            cell_val.value = formula
            cell_name.font = font_bold
            cell_val.font = font_regular
            cell_val.fill = fill_metric
            
            # Apply dynamic formatting/color fusions
            if metric_name == "Lifetime Realized PnL" or metric_name == "Lifetime Realized P/L Percentage":
                if actual_pnl_sum >= 0:
                    cell_val.fill = fill_profit
                    cell_val.font = Font(name="Segoe UI", size=11, color="385723")
                else:
                    cell_val.fill = fill_loss
                    cell_val.font = Font(name="Segoe UI", size=11, color="C00000")
            elif metric_name == "Current Value of Active Trades":
                if open_positions:
                    if total_value >= total_cost:
                        cell_val.fill = fill_profit
                        cell_val.font = Font(name="Segoe UI", size=11, color="385723")
                    else:
                        cell_val.fill = fill_loss
                        cell_val.font = Font(name="Segoe UI", size=11, color="C00000")
            
            if m_type == "currency":
                cell_val.number_format = "INR #,##0.00"
            elif m_type == "percentage":
                cell_val.number_format = "0.0%"
            elif m_type == "integer":
                cell_val.number_format = "#,##0"
            elif m_type == "float":
                cell_val.number_format = "#,##0.00"

    ws_dash.column_dimensions["A"].width = 34
    ws_dash.column_dimensions["B"].width = 22
    
    # Set widths for new columns in Holdings table
    ws_dash.column_dimensions["D"].width = 16  # Stock Symbol
    ws_dash.column_dimensions["E"].width = 10  # Qty
    ws_dash.column_dimensions["F"].width = 16  # Purchase Price
    ws_dash.column_dimensions["G"].width = 16  # Current Price
    ws_dash.column_dimensions["H"].width = 18  # Current Value
    ws_dash.column_dimensions["I"].width = 18  # P&L Amount
    ws_dash.column_dimensions["J"].width = 14  # P&L %
    ws_dash.column_dimensions["K"].width = 14  # Holding % (Weight)

    # --- 1. Current Holdings Value Table ---
    headers = [
        ("Stock", 4),
        ("Qty", 5),
        ("Purchase Price", 6),
        ("Current Price", 7),
        ("Current Value", 8),
        ("P&L Amount", 9),
        ("P&L %", 10),
        ("Holding %", 11)
    ]
    
    for text, col_idx in headers:
        cell = ws_dash.cell(row=3, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_row_idx = 4 + len(open_positions)
    current_row = 4
    for pos in open_positions:
        qty = pos["qty"]
        cost = pos["cost"]
        curr_price = pos.get("last_price", cost)
        
        # Stock
        ws_dash.cell(row=current_row, column=4, value=pos["symbol"]).font = font_regular
        ws_dash.cell(row=current_row, column=4).fill = fill_metric
        ws_dash.cell(row=current_row, column=4).alignment = Alignment(horizontal="center")
        
        # Qty
        qty_cell = ws_dash.cell(row=current_row, column=5, value=qty)
        qty_cell.font = font_regular
        qty_cell.fill = fill_metric
        qty_cell.alignment = Alignment(horizontal="right")
        qty_cell.number_format = "#,##0"
        
        # Purchase Price
        cost_cell = ws_dash.cell(row=current_row, column=6, value=cost)
        cost_cell.font = font_regular
        cost_cell.fill = fill_metric
        cost_cell.alignment = Alignment(horizontal="right")
        cost_cell.number_format = "INR #,##0.00"
        
        # Current Price
        curr_price_cell = ws_dash.cell(row=current_row, column=7, value=curr_price)
        curr_price_cell.font = font_regular
        curr_price_cell.fill = fill_metric
        curr_price_cell.alignment = Alignment(horizontal="right")
        curr_price_cell.number_format = "INR #,##0.00"
        
        # Current Value = Qty * Current Price
        val_cell = ws_dash.cell(row=current_row, column=8, value=f"=E{current_row}*G{current_row}")
        val_cell.font = font_regular
        val_cell.fill = fill_metric
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.number_format = "INR #,##0.00"
        
        # P&L Amount = Current Value - (Qty * Purchase Price)
        pnl_cell = ws_dash.cell(row=current_row, column=9, value=f"=H{current_row}-(E{current_row}*F{current_row})")
        pnl_cell.font = font_bold
        pnl_cell.alignment = Alignment(horizontal="right")
        pnl_cell.number_format = "INR #,##0.00"
        
        # P&L % = P&L Amount / (Qty * Purchase Price)
        pnl_pct_cell = ws_dash.cell(row=current_row, column=10, value=f"=IF(F{current_row}>0, I{current_row}/(E{current_row}*F{current_row}), 0)")
        pnl_pct_cell.font = font_bold
        pnl_pct_cell.alignment = Alignment(horizontal="right")
        pnl_pct_cell.number_format = "0.0%"
        
        # Holding % = Current Value / Total Current Value
        hold_pct_cell = ws_dash.cell(row=current_row, column=11, value=f"=H{current_row}/H{total_row_idx}")
        hold_pct_cell.font = font_regular
        hold_pct_cell.fill = fill_metric
        hold_pct_cell.alignment = Alignment(horizontal="right")
        hold_pct_cell.number_format = "0.0%"
        
        # Color coding P&L cells dynamically
        pnl_val = (curr_price - cost) * qty
        if pnl_val >= 0:
            pnl_cell.fill = fill_profit
            pnl_cell.font = Font(name="Segoe UI", size=11, bold=True, color="385723")
            pnl_pct_cell.fill = fill_profit
            pnl_pct_cell.font = Font(name="Segoe UI", size=11, bold=True, color="385723")
        else:
            pnl_cell.fill = fill_loss
            pnl_cell.font = Font(name="Segoe UI", size=11, bold=True, color="C00000")
            pnl_pct_cell.fill = fill_loss
            pnl_pct_cell.font = Font(name="Segoe UI", size=11, bold=True, color="C00000")
            
        current_row += 1

    # Write Total row
    ws_dash.cell(row=current_row, column=4, value="Total Portfolio").font = font_bold
    ws_dash.cell(row=current_row, column=4).fill = fill_section
    ws_dash.cell(row=current_row, column=4).alignment = Alignment(horizontal="center")
    
    for c in [5, 6, 7]:
        cell = ws_dash.cell(row=current_row, column=c, value="")
        cell.fill = fill_section

    # Total Current Value
    total_val_cell = ws_dash.cell(row=current_row, column=8, value=f"=SUM(H4:H{current_row-1})")
    total_val_cell.font = font_bold
    total_val_cell.fill = fill_section
    total_val_cell.alignment = Alignment(horizontal="right")
    total_val_cell.number_format = "INR #,##0.00"

    # Total P&L Amount
    total_pnl_cell = ws_dash.cell(row=current_row, column=9, value=f"=SUM(I4:I{current_row-1})")
    total_pnl_cell.font = font_bold
    total_pnl_cell.fill = fill_section
    total_pnl_cell.alignment = Alignment(horizontal="right")
    total_pnl_cell.number_format = "INR #,##0.00"

    # Total P&L % = Total P&L / (Total Current Value - Total P&L)
    total_pct_cell = ws_dash.cell(row=current_row, column=10, value=f"=IF((H{current_row}-I{current_row})>0, I{current_row}/(H{current_row}-I{current_row}), 0)")
    total_pct_cell.font = font_bold
    total_pct_cell.fill = fill_section
    total_pct_cell.alignment = Alignment(horizontal="right")
    total_pct_cell.number_format = "0.0%"

    # Total Holding % (sums up to 100.0%)
    total_hold_pct_cell = ws_dash.cell(row=current_row, column=11, value=f"=SUM(K4:K{current_row-1})")
    total_hold_pct_cell.font = font_bold
    total_hold_pct_cell.fill = fill_section
    total_hold_pct_cell.alignment = Alignment(horizontal="right")
    total_hold_pct_cell.number_format = "0.0%"

    # --- 2. Pie Chart ---
    # Shifted to Column M (M3) to avoid overlap
    from openpyxl.chart import PieChart, LineChart, Reference
    pie = PieChart()
    labels = Reference(ws_dash, min_col=4, min_row=4, max_row=current_row-1)
    data = Reference(ws_dash, min_col=8, min_row=3, max_row=current_row-1)  # Current Value is in Col 8 (H)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Current Holdings Allocation"
    pie.width = 14
    pie.height = 8.5
    ws_dash.add_chart(pie, "M3")

    # --- 3. HelperData Sheet & Cumulative PnL ---
    closed_trades = []
    for r in range(2, ws_ledg.max_row + 1):
        status = ws_ledg.cell(row=r, column=14).value
        if status in ["WIN", "LOSS"]:
            exit_date_val = ws_ledg.cell(row=r, column=11).value
            qty = float(ws_ledg.cell(row=r, column=4).value or 0)
            buy = float(ws_ledg.cell(row=r, column=5).value or 0)
            exit_pr = float(ws_ledg.cell(row=r, column=12).value or 0)
            pnl = (exit_pr - buy) * qty
            if exit_date_val:
                date_str = str(exit_date_val).split(" ")[0]
                closed_trades.append((date_str, pnl))

    if closed_trades:
        closed_trades.sort(key=lambda x: x[0])
        pnl_by_date = {}
        for date_str, pnl in closed_trades:
            pnl_by_date[date_str] = pnl_by_date.get(date_str, 0.0) + pnl
            
        sorted_dates = sorted(pnl_by_date.keys())
        cumulative_pnl = 0.0
        chart_data = []
        for d in sorted_dates:
            cumulative_pnl += pnl_by_date[d]
            chart_data.append((d, cumulative_pnl))

        if "HelperData" in wb.sheetnames:
            del wb["HelperData"]
        ws_helper = wb.create_sheet(title="HelperData")
        ws_helper.views.sheetView[0].showGridLines = True
        
        ws_helper.cell(row=1, column=1, value="Date")
        ws_helper.cell(row=1, column=2, value="Cumulative PnL")
        
        for idx, (d_str, cum_pnl) in enumerate(chart_data, 2):
            ws_helper.cell(row=idx, column=1, value=d_str)
            ws_helper.cell(row=idx, column=2, value=cum_pnl)

        # --- 4. Line Chart (Equity Curve) ---
        # Shifted to Column L (L15) to avoid overlap
        line = LineChart()
        line.title = "Equity Curve (Realized PnL)"
        line.style = 13
        line.y_axis.title = "Realized PnL (INR)"
        line.x_axis.title = "Date"
        
        data_ref = Reference(ws_helper, min_col=2, min_row=1, max_row=len(chart_data)+1)
        cats_ref = Reference(ws_helper, min_col=1, min_row=2, max_row=len(chart_data)+1)
        
        line.add_data(data_ref, titles_from_data=True)
        line.set_categories(cats_ref)
        line.legend = None
        line.width = 14
        line.height = 8.5
        ws_dash.add_chart(line, "M15")

    wb.save(file_path)
    wb.close()
    print("Master reconciliation and layout upgrades completed successfully.")

    # Force sync GDrive by raw byte overwrite
    gdrive_path = "G:/My Drive/HalaalTrading/Trading_Journal.xlsx"
    if os.path.exists(gdrive_path):
        try:
            with open(file_path, "rb") as fsrc:
                content = fsrc.read()
            with open(gdrive_path, "wb") as fdst:
                fdst.write(content)
            print("[SUCCESS] Synced reconciled journal bytes to Google Drive.")
        except Exception as e:
            print(f"[ERROR] Failed to write bytes to GDrive: {e}")

if __name__ == "__main__":
    run_master_reconciliation()

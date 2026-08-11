import os
import sys
import datetime
import json
import logging
import openpyxl
import pandas as pd
import concurrent.futures
import yfinance as yf

from hsts.utils import setup_logging
from hsts.sharia import ShariaScreeningEngine
from hsts.regime import MarketRegimeEngine
from hsts.scanner import TechnicalScanner
from hsts.risk import RiskManagementEngine
from hsts.journal import TradingJournal
from hsts.intraday_scanner import HalalIntradayScanner

def check_recommendation_exists(file_path, symbol, date_str, category):
    """
    Check if a recommendation already exists for this symbol, date, and category in the Excel sheet.
    """
    if not os.path.exists(file_path):
        return False
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if "Recommendations" not in wb.sheetnames:
            wb.close()
            return False
        ws = wb["Recommendations"]
        
        # Determine max row
        max_row = ws.max_row
        for r in range(max_row, 0, -1):
            val = ws.cell(row=r, column=1).value
            if val is not None and str(val).strip() != "":
                max_row = r
                break

        for r in range(2, max_row + 1):
            d_val = ws.cell(row=r, column=1).value
            s_val = ws.cell(row=r, column=2).value
            n_val = ws.cell(row=r, column=12).value
            
            if d_val is not None:
                if isinstance(d_val, (datetime.date, datetime.datetime)):
                    d_str = d_val.strftime("%Y-%m-%d")
                else:
                    d_str = str(d_val).split(" ")[0].strip()
            else:
                d_str = ""
            
            if d_str == date_str and s_val == symbol and n_val and category in str(n_val):
                wb.close()
                return True
        wb.close()
        return False
    except Exception as e:
        print(f"Error checking recommendation: {e}")
        return False

def run_daily_scan():
    setup_logging()
    logger = logging.getLogger("hsts.daily_scan")
    logger.info("Initializing HSTS Daily Scan Orchestrator...")
    
    universe_path = "data/universe.csv"
    journal_path = "Trading_Journal.xlsx"
    date_str = datetime.date.today().strftime("%Y-%m-%d")
    
    if not os.path.exists(universe_path):
        logger.error(f"Stock universe file not found at {universe_path}")
        sys.exit(1)
        
    sharia_engine = ShariaScreeningEngine(universe_path)
    regime_engine = MarketRegimeEngine()
    scanner = TechnicalScanner()
    risk_engine = RiskManagementEngine()
    journal = TradingJournal(file_path=journal_path)
    intraday_scanner = HalalIntradayScanner(universe_csv_path=universe_path)
    
    capital = journal.get_available_capital()
    regime, regime_metrics = regime_engine.get_market_regime()
    alloc_cap_pct = regime_metrics.get("allocation_cap", 0.20)
    
    logger.info(f"Market Regime: {regime} (Allocation Cap: {alloc_cap_pct*100:.0f}%)")
    logger.info(f"Available Capital: INR {capital:,.2f}")
    
    # ---------------------------------------------
    # 1. SWING TRADING SCAN
    # ---------------------------------------------
    logger.info("Running Swing Trading scan...")
    df_universe = pd.read_csv(universe_path)
    compliant_symbols = []
    
    for idx, row in df_universe.iterrows():
        symbol = row["symbol"]
        name = row["name"]
        is_halal, screen_details = sharia_engine.screen_stock(symbol)
        if is_halal:
            compliant_symbols.append((symbol, name))
            
    swing_setups = []
    if compliant_symbols:
        bad_symbols = {"LTIM", "BIRLASOFT", "OBEROIREAL", "GUJGASLTD", "JBCHEPHARM", "SPICEJET", "CSM"}
        filtered_compliant = [(sym, name) for sym, name in compliant_symbols if sym not in bad_symbols]
        
        tickers_map = {f"{sym}.NS": (sym, name) for sym, name in filtered_compliant}
        tickers_list = list(tickers_map.keys())
        
        # Batch download prices
        batch_size = 100
        chunks = [tickers_list[i:i+batch_size] for i in range(0, len(tickers_list), batch_size)]
        
        dfs = []
        for index, chunk in enumerate(chunks):
            try:
                df = yf.download(chunk, period="6mo", group_by="ticker", progress=False, threads=True, timeout=12)
                if not df.empty:
                    dfs.append(df)
            except Exception as ex:
                logger.warning(f"Failed to download batch {index+1}: {ex}")
                
        if dfs:
            df_batch = pd.concat(dfs, axis=1)
        else:
            df_batch = pd.DataFrame()
            
        all_compliant_analyses = []
        for ns_sym, (symbol, name) in tickers_map.items():
            df_ticker = None
            if len(tickers_list) == 1:
                df_ticker = df_batch.dropna(subset=["Close"])
            elif ns_sym in df_batch.columns.levels[0]:
                df_ticker = df_batch[ns_sym].dropna(subset=["Close"])
                
            if df_ticker is None or df_ticker.empty or len(df_ticker) < 50:
                continue
                
            analysis = scanner.analyze_stock(symbol, df=df_ticker)
            if "reason" in analysis and analysis["signal"] == "WAIT" and "No historical" in analysis["reason"]:
                continue
                
            all_compliant_analyses.append({
                "symbol": symbol,
                "name": name,
                "close": analysis["close"],
                "rsi": analysis["rsi"],
                "score": analysis["score"],
                "signal": analysis["signal"],
                "suggested_sl": analysis["suggested_sl"],
                "suggested_target": analysis["suggested_target"]
            })
            
        if all_compliant_analyses:
            df_all = pd.DataFrame(all_compliant_analyses)
            # Sort by score descending and take top 5
            df_top5 = df_all.sort_values(by="score", ascending=False).head(5)
            
            for _, row in df_top5.iterrows():
                pos_size = risk_engine.calculate_position_size(
                    total_capital=capital,
                    entry_price=row["close"],
                    stop_loss_price=row["suggested_sl"],
                    max_allocation_pct=alloc_cap_pct
                )
                
                risk = row["close"] - row["suggested_sl"]
                reward = row["suggested_target"] - row["close"]
                rr_val = (reward / risk) if risk > 0 else 2.0
                
                setup = {
                    "symbol": row["symbol"],
                    "name": row["name"],
                    "close": float(row["close"]),
                    "score": int(row["score"]),
                    "signal": row["signal"],
                    "stop_loss": float(row["suggested_sl"]),
                    "target": float(row["suggested_target"]),
                    "qty": int(pos_size["quantity"]) if pos_size else 0,
                    "allocation": float(pos_size["total_investment"]) if pos_size else 0.0,
                    "rr_ratio": f"1:{rr_val:.1f}"
                }
                swing_setups.append(setup)
                
                # Check for duplicate and log to Excel
                if not check_recommendation_exists(journal_path, row["symbol"], date_str, "Swing Scanner"):
                    exec_status = "SKIPPED_BEARISH_REGIME" if regime == "STATE 4: Cash Only" else "PENDING"
                    journal.add_recommendation(
                        symbol=row["symbol"],
                        name=row["name"],
                        score=row["score"],
                        target_entry=row["close"],
                        stop_loss=row["suggested_sl"],
                        profit_target=row["suggested_target"],
                        qty=setup["qty"],
                        allocation=setup["allocation"],
                        status=exec_status,
                        notes=f"Swing Scanner ({regime})"
                    )
                    
    # ---------------------------------------------
    # 2. INTRADAY TRADING SCAN
    # ---------------------------------------------
    logger.info("Running Intraday Trading scan...")
    intraday_candidates = intraday_scanner.scan_universe()
    intraday_setups = []
    
    if intraday_candidates:
        df_intra = pd.DataFrame(intraday_candidates)
        # Sort by score descending and take top 5
        df_intra_top5 = df_intra.sort_values(by="score", ascending=False).head(5)
        
        for _, row in df_intra_top5.iterrows():
            setup = {
                "symbol": row["symbol"],
                "score": int(row["score"]),
                "entry": float(row["entry"]),
                "stop_loss": float(row["stop_loss"]),
                "target": float(row["target"]),
                "rr_ratio": f"1:{row['rr_ratio']:.1f}"
            }
            intraday_setups.append(setup)
            
            # Fetch company name from universe
            matched_univ = df_universe[df_universe["symbol"] == row["symbol"]]
            comp_name = matched_univ.iloc[0]["name"] if not matched_univ.empty else "Unknown Name"
            
            # Check for duplicate and log to Excel Recommendations
            if not check_recommendation_exists(journal_path, row["symbol"], date_str, "Intraday Scanner"):
                # Calculate estimated quantity using 10% allocation cap as standard for Intraday
                pos_size_intra = risk_engine.calculate_position_size(
                    total_capital=capital,
                    entry_price=row["entry"],
                    stop_loss_price=row["stop_loss"],
                    max_allocation_pct=0.10
                )
                qty_intra = pos_size_intra["quantity"] if pos_size_intra else 0
                alloc_intra = pos_size_intra["total_investment"] if pos_size_intra else 0.0
                
                journal.add_recommendation(
                    symbol=row["symbol"],
                    name=comp_name,
                    score=row["score"],
                    target_entry=row["entry"],
                    stop_loss=row["stop_loss"],
                    profit_target=row["target"],
                    qty=qty_intra,
                    allocation=alloc_intra,
                    status="PENDING",
                    notes=f"Intraday Scanner ({regime})"
                )

    # ---------------------------------------------
    # 3. GENERATE OUTPUTS
    # ---------------------------------------------
    logger.info("Generating reports...")
    
    # Save JSON data
    output_data = {
        "date": date_str,
        "market_regime": regime,
        "swing_recommendations": swing_setups,
        "intraday_recommendations": intraday_setups
    }
    
    os.makedirs("data", exist_ok=True)
    with open("data/daily_recommendations.json", "w", encoding="utf-8") as f:
        json.dump(output_data, f, indent=2)
        
    # Generate Markdown Report
    report_content = f"""# Daily Trading Recommendations & Scanner Report

*Generated on: **{date_str}***
*Broader Market Regime: **{regime}** (Allocation Cap: **{alloc_cap_pct*100:.0f}%**)*

---

## 📈 Swing Trading Recommendations (Top 5 Compliant)
Swing positions are momentum setups with holding periods ranging from a few days to several weeks.

"""
    if swing_setups:
        report_content += "| Symbol | Company Name | Momentum Score | Signal | Entry Price | Stop Loss | Profit Target | Risk/Reward | Est. Qty | Est. Allocation |\n"
        report_content += "| :--- | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :---: |\n"
        for s in swing_setups:
            report_content += f"| **{s['symbol']}** | {s['name']} | {s['score']}/100 | `{s['signal']}` | INR {s['close']:.2f} | INR {s['stop_loss']:.2f} | INR {s['target']:.2f} | {s['rr_ratio']} | {s['qty']} | INR {s['allocation']:,.2f} |\n"
        
        if regime == "STATE 4: Cash Only":
            report_content += "\n> [Slim Warning]\n> **BROADER MARKET IS IN CASH ONLY REGIME.**\n> Under HSTS rules, taking new swing momentum entries is strictly not recommended to preserve capital.\n"
    else:
        report_content += "*No compliant swing setups found meeting the score threshold today.*\n"

    report_content += """
---

## ⚡ Intraday Trading Recommendations (Top 5 Compliant)
Intraday positions are breakout setups that must be closed before 3:15 PM IST.

"""
    if intraday_setups:
        report_content += "| Symbol | Composite Score | Target Entry | Stop Loss | Profit Target | Risk/Reward |\n"
        report_content += "| :--- | :---: | :---: | :---: | :---: | :---: |\n"
        for s in intraday_setups:
            report_content += f"| **{s['symbol']}** | {s['score']}/100 | INR {s['entry']:.2f} | INR {s['stop_loss']:.2f} | INR {s['target']:.2f} | {s['rr_ratio']} |\n"
            
        report_content += """
> [Slim Important]
> **AAOIFI Shariah Compliance Rules for Intraday Trading:**
> 1. All intraday orders **must be placed as CNC (Cash Delivery)** on your broker platform.
> 2. The use of MIS, intraday margin, or leverage is strictly prohibited.
> 3. Any position that does not hit its profit target or stop loss must be **squared off manually before 3:15 PM IST**.
"""
    else:
        report_content += "*No breakout setups meeting the criteria were identified today.*\n"

    report_content += """
---
*Disclaimer: All recommendations are generated by the rule-based Halal Swing & Trend System (HSTS) scanning model. Please perform your own diligence before executing trades.*
"""

    with open("Daily_Recommendations.md", "w", encoding="utf-8") as f:
        f.write(report_content)
        
    logger.info("Daily scanning and report generation completed successfully!")

if __name__ == "__main__":
    run_daily_scan()

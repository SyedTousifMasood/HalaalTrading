import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import logging

logger = logging.getLogger("hsts.journal")

class TradingJournal:
    def __init__(self, file_path="Trading_Journal.xlsx"):
        self.file_path = file_path
        self.initialize_journal()

    def _get_true_max_row(self, ws):
        """Find the actual last populated row in Column A to prevent writing 1000 rows down."""
        for r in range(ws.max_row, 0, -1):
            val = ws.cell(row=r, column=1).value
            if val is not None and str(val).strip() != "":
                return r
        return 1

    def initialize_journal(self):
        """
        Creates, formats, or updates the Excel Workbook.
        """
        if os.path.exists(self.file_path):
            logger.info("Trading Journal already exists. Checking for updates/missing sheets...")
            wb = openpyxl.load_workbook(self.file_path)
            
            # Upgrade check: Add Capital sheet if missing
            if "Capital" not in wb.sheetnames:
                logger.info("Upgrading Trading Journal: Adding Capital sheet...")
                ws_cap = wb.create_sheet(title="Capital")
                self._setup_capital(ws_cap)

            # Upgrade check: Add Recommendations sheet if missing
            if "Recommendations" not in wb.sheetnames:
                logger.info("Upgrading Trading Journal: Adding Recommendations sheet...")
                ws_recs = wb.create_sheet(title="Recommendations")
                self._setup_recommendations(ws_recs)

            ws_dash = wb["Dashboard"]
            self._setup_dashboard(ws_dash)

            wb.save(self.file_path)
            wb.close()
            self._upgrade_rr_ratio_columns()
            return

        logger.info(f"Creating a new Trading Journal at {self.file_path}...")
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
        
        wb = openpyxl.Workbook()
        
        # 1. Setup Dashboard Sheet
        ws_dash = wb.active
        ws_dash.title = "Dashboard"
        self._setup_dashboard(ws_dash)

        # 2. Setup Ledger Sheet
        ws_ledg = wb.create_sheet(title="Ledger")
        self._setup_ledger(ws_ledg)

        # 3. Setup Recommendations Sheet
        ws_recs = wb.create_sheet(title="Recommendations")
        self._setup_recommendations(ws_recs)

        # 4. Setup Capital Sheet
        ws_cap = wb.create_sheet(title="Capital")
        self._setup_capital(ws_cap)

        # 5. Setup Logs Sheet
        ws_logs = wb.create_sheet(title="Logs")
        self._setup_logs(ws_logs)

        wb.save(self.file_path)
        logger.info("Workbook created and formatted successfully.")

    def _upgrade_rr_ratio_columns(self):
        wb = openpyxl.load_workbook(self.file_path)
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header_recs = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
        fill_header_ledg = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")

        # 1. Recommendations Sheet Upgrade
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            if ws.cell(row=1, column=8).value != "Risk-to-Reward Ratio":
                logger.info("Upgrading Recommendations sheet: Adding Risk-to-Reward Ratio column...")
                ws.insert_cols(8)
                cell = ws.cell(row=1, column=8, value="Risk-to-Reward Ratio")
                cell.font = font_header
                cell.fill = fill_header_recs
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions["H"].width = 22

                true_max = self._get_true_max_row(ws)
                for r in range(2, true_max + 1):
                    entry = ws.cell(row=r, column=5).value
                    sl = ws.cell(row=r, column=6).value
                    target = ws.cell(row=r, column=7).value
                    
                    if isinstance(entry, (int, float)) and isinstance(sl, (int, float)) and isinstance(target, (int, float)):
                        risk = entry - sl
                        reward = target - entry
                        rr_val = (reward / risk) if risk > 0 else 2.0
                        ws.cell(row=r, column=8, value=f"1:{rr_val:.1f}")
                    else:
                        ws.cell(row=r, column=8, value="1:2.0")

        # 2. Ledger Sheet Upgrade
        if "Ledger" in wb.sheetnames:
            ws = wb["Ledger"]
            if ws.cell(row=1, column=10).value != "Risk-to-Reward Ratio":
                logger.info("Upgrading Ledger sheet: Adding Risk-to-Reward Ratio column...")
                ws.insert_cols(10)
                cell = ws.cell(row=1, column=10, value="Risk-to-Reward Ratio")
                cell.font = font_header
                cell.fill = fill_header_ledg
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions["J"].width = 22

                true_max = self._get_true_max_row(ws)
                for r in range(2, true_max + 1):
                    buy_price = ws.cell(row=r, column=5).value
                    target = ws.cell(row=r, column=8).value
                    sl = ws.cell(row=r, column=9).value
                    
                    if isinstance(buy_price, (int, float)) and isinstance(sl, (int, float)) and isinstance(target, (int, float)):
                        risk = buy_price - sl
                        reward = target - buy_price
                        rr_val = (reward / risk) if risk > 0 else 2.0
                        ws.cell(row=r, column=10, value=f"1:{rr_val:.1f}")
                    else:
                        ws.cell(row=r, column=10, value="1:2.0")
                    
                    pnl_formula = f'=IF(OR(N{r}="WIN", N{r}="LOSS"), (L{r}-E{r})*D{r}, 0)'
                    ws.cell(row=r, column=13, value=pnl_formula)

            # 3. Trade Type Column P Upgrade
            if ws.cell(row=1, column=16).value != "Trade Type":
                logger.info("Upgrading Ledger sheet: Adding Trade Type column...")
                cell = ws.cell(row=1, column=16, value="Trade Type")
                cell.font = font_header
                cell.fill = fill_header_ledg
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions["P"].width = 18

                true_max = self._get_true_max_row(ws)
                for r in range(2, true_max + 1):
                    val = ws.cell(row=r, column=16).value
                    if val is None or str(val).strip() == "":
                        ws.cell(row=r, column=16, value="SWING")

        ws_dash = wb["Dashboard"]
        self._setup_dashboard(ws_dash)

        wb.save(self.file_path)

    def _setup_dashboard(self, ws):
        ws.views.sheetView[0].showGridLines = True
        font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Segoe UI", size=11, bold=True)
        font_regular = Font(name="Segoe UI", size=11)
        font_section = Font(name="Segoe UI", size=11, bold=True, color="1B365D")
        fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        fill_metric = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
        fill_section = PatternFill(start_color="D3DDF1", end_color="D3DDF1", fill_type="solid")

        ws["A1"] = "HSTS v1.0 Trading Dashboard"
        ws["A1"].font = font_title
        ws.row_dimensions[1].height = 30

        existing_a2 = ws["A2"].value
        if existing_a2:
            ws["A2"] = existing_a2
        else:
            ws["A2"] = f"Last updated on : {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(name="Segoe UI", size=9, italic=True, color="595959")
        ws.row_dimensions[2].height = 18

        headers = ["Metric", "Value"]
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 24

        # Unmerge all cells first to prevent layout collisions during template rebuild
        for rng in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(rng))

        # Try to read current value of active trades to avoid resetting it to 0.0
        current_b12 = ws.cell(row=12, column=2).value
        try:
            curr_val_active = float(current_b12) if current_b12 is not None else 0.0
        except ValueError:
            curr_val_active = 0.0

        # Calculate analytics from Ledger sheet
        ws_ledg = ws.parent["Ledger"]
        closed_durations = []
        entry_dates = set()
        trades_with_dates = []
        min_date = None
        max_date = datetime.date.today()
        
        for r in range(2, ws_ledg.max_row + 1):
            sym = ws_ledg.cell(row=r, column=1).value
            status = ws_ledg.cell(row=r, column=14).value
            entry_val = ws_ledg.cell(row=r, column=3).value
            exit_val = ws_ledg.cell(row=r, column=11).value
            
            if sym and entry_val:
                try:
                    if isinstance(entry_val, (datetime.datetime, datetime.date)):
                        ent_d = entry_val.date() if isinstance(entry_val, datetime.datetime) else entry_val
                    else:
                        ent_d = datetime.datetime.strptime(str(entry_val).split(" ")[0], "%Y-%m-%d").date()
                    
                    entry_dates.add(ent_d)
                    
                    ext_d = None
                    if exit_val:
                        if isinstance(exit_val, (datetime.datetime, datetime.date)):
                            ext_d = exit_val.date() if isinstance(exit_val, datetime.datetime) else exit_val
                        else:
                            ext_d = datetime.datetime.strptime(str(exit_val).split(" ")[0], "%Y-%m-%d").date()
                    
                    trades_with_dates.append((sym, ent_d, ext_d))
                    
                    if min_date is None or ent_d < min_date:
                        min_date = ent_d
                        
                    if status in ["WIN", "LOSS"] and ext_d:
                        closed_durations.append((ext_d - ent_d).days)
                    elif status == "OPEN":
                        closed_durations.append((max_date - ent_d).days)
                except Exception:
                    pass
                    
        avg_exit_days = sum(closed_durations) / len(closed_durations) if closed_durations else 0.0
        avg_trades_per_day = len(trades_with_dates) / len(entry_dates) if entry_dates else 0.0
        
        daily_counts = []
        if min_date:
            d = min_date
            while d <= max_date:
                if d.weekday() < 5:
                    count = sum(1 for sym, ent_d, ext_d in trades_with_dates if ent_d <= d <= (ext_d if ext_d else max_date))
                    daily_counts.append(count)
                d += datetime.timedelta(days=1)
        avg_unique_holdings = sum(daily_counts) / len(daily_counts) if daily_counts else 0.0

        metrics = [
            ("Global Portfolio Metrics", None, "header"),
            ("Total Net Capital Deposited", '=SUMIF(Capital!B:B, "DEPOSIT", Capital!C:C)', "currency"),
            ("Lifetime Realized PnL", "=SUM(Ledger!M:M)", "currency"),
            ("Lifetime Realized P/L Percentage", "=B6/B5", "percentage"),
            ("Current Total Portfolio Value", "=B13+B12", "currency"),
            ("Total Number of Trades", '=COUNTIF(Ledger!N:N, "WIN") + COUNTIF(Ledger!N:N, "LOSS") + COUNTIF(Ledger!N:N, "OPEN")', "integer"),
            ("Winning Percentage of Trades", '=IF((COUNTIF(Ledger!N:N, "WIN")+COUNTIF(Ledger!N:N, "LOSS"))>0, COUNTIF(Ledger!N:N, "WIN")/(COUNTIF(Ledger!N:N, "WIN")+COUNTIF(Ledger!N:N, "LOSS")), 0)', "percentage"),
            ("Capital Deployed in Open Trades", '=SUMPRODUCT((Ledger!N2:N500="OPEN")*(Ledger!E2:E500)*(Ledger!D2:D500))', "currency"),
            ("Current Value of Active Trades", curr_val_active, "currency"),
            ("Capital Available for Trading", "=B5 - SUMIF(Capital!B:B, \"WITHDRAWAL\", Capital!C:C) + B6 - B11", "currency"),
            ("Total Wins", '=COUNTIF(Ledger!N:N, "WIN")', "integer"),
            ("Total Losses", '=COUNTIF(Ledger!N:N, "LOSS")', "integer"),
            ("Total Taxes Paid to Zerodha", '=SUMIF(Capital!D:D, "*Brokerage & taxes*", Capital!C:C)', "currency"),
            ("Swing Trading Statistics", None, "header"),
            ("Swing Win Rate", '=IF((COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "WIN")+COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "LOSS"))>0, COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "WIN")/(COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "WIN")+COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "LOSS")), 0)', "percentage"),
            ("Swing Realized PnL", '=SUMIFS(Ledger!M:M, Ledger!P:P, "SWING")', "currency"),
            ("Swing Completed Trades", '=COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "WIN") + COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "LOSS")', "integer"),
            ("Swing Active Positions", '=COUNTIFS(Ledger!P:P, "SWING", Ledger!N:N, "OPEN")', "integer"),
            ("Intraday Trading Statistics", None, "header"),
            ("Intraday Win Rate", '=IF((COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "WIN")+COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "LOSS"))>0, COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "WIN")/(COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "WIN")+COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "LOSS")), 0)', "percentage"),
            ("Intraday Realized PnL", '=SUMIFS(Ledger!M:M, Ledger!P:P, "INTRADAY")', "currency"),
            ("Intraday Completed Trades", '=COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "WIN") + COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "LOSS")', "integer"),
            ("Intraday Active Positions", '=COUNTIFS(Ledger!P:P, "INTRADAY", Ledger!N:N, "OPEN")', "integer"),
            ("Activity & Performance Metrics", None, "header"),
            ("Average Days to Exit a Trade", avg_exit_days, "float"),
            ("Average Trades per Day", avg_trades_per_day, "float"),
            ("Average Unique Holdings per Day", avg_unique_holdings, "float"),
        ]

        # Colors for dynamic active trades value
        fill_profit = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") # Pastel Green
        fill_loss = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Pastel Red

        for idx, (metric_name, formula, m_type) in enumerate(metrics, 4):
            cell_name = ws.cell(row=idx, column=1, value=metric_name)
            cell_val = ws.cell(row=idx, column=2)
            
            if m_type == "header":
                cell_name.font = font_section
                cell_name.fill = fill_section
                cell_val.value = ""
                cell_val.fill = fill_section
                # Merge columns A and B for header row
                ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=2)
            else:
                cell_val.value = formula
                cell_name.font = font_bold
                cell_val.font = font_regular
                cell_val.fill = fill_metric
                
                if m_type == "currency":
                    cell_val.number_format = "INR #,##0.00"
                elif m_type == "percentage":
                    cell_val.number_format = "0.0%"
                elif m_type == "integer":
                    cell_val.number_format = "#,##0"
                elif m_type == "float":
                    cell_val.number_format = "#,##0.00"
                elif m_type == "integer":
                    cell_val.number_format = "#,##0"

        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 22

    def _setup_ledger(self, ws):
        ws.views.sheetView[0].showGridLines = True
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")

        headers = [
            "Symbol", "Name", "Entry Date", "Qty", "Buy Price", 
            "Suggested Entry", "Slippage/Deviation", "Target Price", 
            "Stop Loss", "Risk-to-Reward Ratio", "Exit Date", "Exit Price", 
            "Realized PnL", "Status", "Notes", "Trade Type"
        ]

        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.row_dimensions[1].height = 28

        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 16
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["J"].width = 22
        ws.column_dimensions["O"].width = 30
        ws.column_dimensions["P"].width = 18

    def _setup_recommendations(self, ws):
        ws.views.sheetView[0].showGridLines = True
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")

        headers = [
            "Date", "Symbol", "Name", "Composite Score", 
            "Target Entry Price", "Initial Stop-Loss", "Profit Target", 
            "Risk-to-Reward Ratio", "Recommended Qty", "Recommended Allocation", 
            "Execution Status", "Notes"
        ]

        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 28

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 22
        ws.column_dimensions["I"].width = 18
        ws.column_dimensions["J"].width = 22
        ws.column_dimensions["K"].width = 24
        ws.column_dimensions["L"].width = 30

    def get_available_capital(self, default_capital=110000.0):
        """Calculates Capital Available for Trading dynamically from the Journal."""
        if not os.path.exists(self.file_path):
            return default_capital
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws_cap = wb["Capital"]
            deposits = 0.0
            has_transactions = False
            for r in range(2, self._get_true_max_row(ws_cap) + 1):
                t_type = str(ws_cap.cell(row=r, column=2).value or "").strip().upper()
                amt = float(ws_cap.cell(row=r, column=3).value or 0.0)
                if t_type in ["DEPOSIT", "ADJUSTMENT_IN"]:
                    deposits += amt
                    has_transactions = True
                elif t_type in ["WITHDRAWAL", "ADJUSTMENT_OUT"]:
                    deposits -= amt
                    has_transactions = True
            if not has_transactions:
                deposits = default_capital

            ws_ledg = wb["Ledger"]
            deployed = 0.0
            realized_pnl = 0.0
            for r in range(2, self._get_true_max_row(ws_ledg) + 1):
                status = str(ws_ledg.cell(row=r, column=14).value or "").strip().upper()
                qty = float(ws_ledg.cell(row=r, column=4).value or 0.0)
                price = float(ws_ledg.cell(row=r, column=5).value or 0.0)
                
                # Column 13 is Realized PnL (Formula or float)
                pnl_val = ws_ledg.cell(row=r, column=13).value
                if isinstance(pnl_val, (int, float)):
                    pnl = float(pnl_val)
                else:
                    # Calculate manually if formula string
                    exit_p = float(ws_ledg.cell(row=r, column=12).value or price)
                    pnl = (exit_p - price) * qty

                if status == "OPEN":
                    deployed += (qty * price)
                elif status in ["CLOSED", "WIN", "LOSS"]:
                    realized_pnl += pnl

            wb.close()
            avail = deposits + realized_pnl - deployed
            return avail if avail > 0 else default_capital
        except Exception as e:
            logger.error(f"Error reading available capital from journal: {e}")
            return default_capital

    def _setup_capital(self, ws):
        ws.views.sheetView[0].showGridLines = True
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")

        headers = ["Date", "Type", "Amount", "Notes"]
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 40

    def _setup_logs(self, ws):
        ws.views.sheetView[0].showGridLines = True
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")

        headers = ["Timestamp", "Level", "Message"]
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 80

    def add_recommendation(self, symbol, name, score, target_entry, stop_loss, profit_target, qty, allocation, status="PENDING", notes=""):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Recommendations"]
        row_idx = self._get_true_max_row(ws) + 1
        
        date_str = datetime.date.today().strftime("%Y-%m-%d")

        risk = target_entry - stop_loss
        reward = profit_target - target_entry
        rr_val = (reward / risk) if risk > 0 else 2.0
        rr_ratio_str = f"1:{rr_val:.1f}"

        ws.cell(row=row_idx, column=1, value=date_str)
        ws.cell(row=row_idx, column=2, value=symbol)
        ws.cell(row=row_idx, column=3, value=name)
        ws.cell(row=row_idx, column=4, value=f"{score:.0f}/100")
        ws.cell(row=row_idx, column=5, value=target_entry)
        ws.cell(row=row_idx, column=6, value=stop_loss)
        ws.cell(row=row_idx, column=7, value=profit_target)
        ws.cell(row=row_idx, column=8, value=rr_ratio_str)
        ws.cell(row=row_idx, column=9, value=qty)
        ws.cell(row=row_idx, column=10, value=allocation)
        ws.cell(row=row_idx, column=11, value=status)
        ws.cell(row=row_idx, column=12, value=notes)

        ws.cell(row=row_idx, column=5).number_format = "INR #,##0.00"
        ws.cell(row=row_idx, column=6).number_format = "INR #,##0.00"
        ws.cell(row=row_idx, column=7).number_format = "INR #,##0.00"
        ws.cell(row=row_idx, column=10).number_format = "INR #,##0.00"

        wb.save(self.file_path)
        logger.info(f"Logged recommendation for {symbol} to Recommendations sheet (Row {row_idx})")

    def capital_transaction_exists(self, notes):
        """Checks if a capital transaction with the exact notes already exists to prevent duplicate charges."""
        if not os.path.exists(self.file_path):
            return False
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb["Capital"]
            for r in range(2, self._get_true_max_row(ws) + 1):
                val = ws.cell(row=r, column=4).value
                if val == notes:
                    wb.close()
                    return True
            wb.close()
            return False
        except Exception as e:
            logger.error(f"Error checking capital transaction existence: {e}")
            return False

    def add_capital_transaction(self, transaction_type, amount, notes=""):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Capital"]
        row_idx = self._get_true_max_row(ws) + 1
        
        date_str = datetime.date.today().strftime("%Y-%m-%d")
        ws.cell(row=row_idx, column=1, value=date_str)
        ws.cell(row=row_idx, column=2, value=transaction_type.upper())
        ws.cell(row=row_idx, column=3, value=amount)
        ws.cell(row=row_idx, column=4, value=notes)

        ws.cell(row=row_idx, column=3).number_format = "INR #,##0.00"
        
        wb.save(self.file_path)
        logger.info(f"Capital Transaction: {transaction_type.upper()} of INR {amount:.2f} logged.")
        self.log_event(f"Capital Transaction: {transaction_type.upper()} of INR {amount:.2f} logged.")

    def add_trade(self, symbol, name, entry_date, qty, buy_price, suggested_entry, target, stop_loss, notes="", trade_type="SWING"):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Ledger"]
        row_idx = self._get_true_max_row(ws) + 1
        
        # Prevent duplicate entries for same symbol/date
        for r in range(2, row_idx):
            if ws.cell(row=r, column=1).value == symbol and ws.cell(row=r, column=3).value == entry_date and ws.cell(row=r, column=14).value == "OPEN":
                logger.info(f"Trade for {symbol} on {entry_date} already exists at Row {r}. Skipping duplicate.")
                wb.close()
                return

        slippage = buy_price - suggested_entry
        risk = buy_price - stop_loss
        reward = target - buy_price
        rr_val = (reward / risk) if risk > 0 else 2.0
        rr_ratio_str = f"1:{rr_val:.1f}"

        ws.cell(row=row_idx, column=1, value=symbol)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=entry_date)
        ws.cell(row=row_idx, column=4, value=qty)
        ws.cell(row=row_idx, column=5, value=buy_price)
        ws.cell(row=row_idx, column=6, value=suggested_entry)
        ws.cell(row=row_idx, column=7, value=slippage)
        ws.cell(row=row_idx, column=8, value=target)
        ws.cell(row=row_idx, column=9, value=stop_loss)
        ws.cell(row=row_idx, column=10, value=rr_ratio_str)
        
        ws.cell(row=row_idx, column=11, value="")
        ws.cell(row=row_idx, column=12, value="")
        pnl_formula = f'=IF(OR(N{row_idx}="WIN", N{row_idx}="LOSS"), (L{row_idx}-E{row_idx})*D{row_idx}, 0)'
        ws.cell(row=row_idx, column=13, value=pnl_formula)
        ws.cell(row=row_idx, column=14, value="OPEN")
        ws.cell(row=row_idx, column=15, value=notes)
        ws.cell(row=row_idx, column=16, value=trade_type.upper())

        ws.cell(row=row_idx, column=5).number_format = "INR #,##0.00"
        ws.cell(row=row_idx, column=6).number_format = "INR #,##0.00"
        ws.cell(row=row_idx, column=7).number_format = "INR #,##0.00"
        ws.cell(row=row_idx, column=8).number_format = "INR #,##0.00"
        ws.cell(row=row_idx, column=9).number_format = "INR #,##0.00"
        ws.cell(row=row_idx, column=13).number_format = "INR #,##0.00"

        wb.save(self.file_path)
        wb.close()
        self.apply_row_styling()
        logger.info(f"Recorded open trade for {symbol} to Ledger (Row {row_idx})")
        self.log_event(f"Recorded buy order: {qty} shares of {symbol} at {buy_price} ({trade_type.upper()})")

    def update_open_trade(self, symbol, actual_buy_price, actual_entry_date):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Ledger"]
        true_max = self._get_true_max_row(ws)
        for r in range(2, true_max + 1):
            if ws.cell(row=r, column=1).value == symbol and ws.cell(row=r, column=14).value == "OPEN":
                ws.cell(row=r, column=5, value=actual_buy_price)
                ws.cell(row=r, column=3, value=actual_entry_date)
                wb.save(self.file_path)
                wb.close()
                self.log_event(f"Updated OPEN trade: {symbol} execution price to {actual_buy_price} and date to {actual_entry_date}")
                logger.info(f"Updated OPEN trade: {symbol} execution price to {actual_buy_price} and date to {actual_entry_date}")
                return True
        wb.close()
        return False

    def get_open_trade_buy_price(self, symbol):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Ledger"]
        true_max = self._get_true_max_row(ws)
        for r in range(2, true_max + 1):
            if ws.cell(row=r, column=1).value == symbol and ws.cell(row=r, column=14).value == "OPEN":
                buy_price = ws.cell(row=r, column=5).value
                wb.close()
                return buy_price
        wb.close()
        return None

    def close_trade(self, symbol, exit_date, exit_price, status, notes=""):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Ledger"]
        found = False
        true_max = self._get_true_max_row(ws)
        for r in range(2, true_max + 1):
            if ws.cell(row=r, column=1).value == symbol and ws.cell(row=r, column=14).value == "OPEN":
                ws.cell(row=r, column=11, value=exit_date)
                ws.cell(row=r, column=12, value=exit_price)
                ws.cell(row=r, column=14, value=status.upper())
                if notes:
                    ws.cell(row=r, column=15, value=f"{ws.cell(row=r, column=15).value} | {notes}".strip(" |"))
                ws.cell(row=r, column=12).number_format = "INR #,##0.00"
                found = True
                break

        if not found:
            logger.error(f"No active open trade found for symbol {symbol}")
            wb.close()
            return False

        wb.save(self.file_path)
        wb.close()
        
        self.apply_row_styling()
        self.log_event(f"Closed trade: {symbol} exited at {exit_price} ({status})")
        return True

    def apply_row_styling(self):
        """Color code completed Ledger rows based on Win/Loss status."""
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Ledger"]
        ws.views.sheetView[0].showGridLines = True
        
        fill_win = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") # Pastel Green
        fill_loss = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Pastel Red
        
        true_max = self._get_true_max_row(ws)
        for r in range(2, true_max + 1):
            status = ws.cell(row=r, column=14).value
            if not status:
                continue
            status_upper = str(status).upper()
            if status_upper == "WIN":
                row_fill = fill_win
            elif status_upper == "LOSS":
                row_fill = fill_loss
            else:
                row_fill = None
                
            if row_fill:
                for c in range(1, 17): # Columns A to P
                    ws.cell(row=r, column=c).fill = row_fill
        wb.save(self.file_path)
        wb.close()

    def log_event(self, message, level="INFO"):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Logs"]
        row_idx = self._get_true_max_row(ws) + 1
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=row_idx, column=1, value=timestamp)
        ws.cell(row=row_idx, column=2, value=level.upper())
        ws.cell(row=row_idx, column=3, value=message)
        wb.save(self.file_path)
